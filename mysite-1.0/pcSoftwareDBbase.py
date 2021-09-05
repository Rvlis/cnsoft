#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Project      ：mysite
@File         : loacalSoftware.py
@Time         : 2021/8/28 22:14
@Author       : lgl
@PRODUCT_NAME : PyCharm
"""
import os
import tkinter
import tkinter.filedialog
import connectMySQL
import time
from prediction import prediction_fun
import xlrd
import openpyxl
from pymysql.converters import escape_string

global forecastdataLen
forecastdataLen = connectMySQL.forecastdataLen()

import socket

# 创建窗口
win = tkinter.Tk()
win.title('新闻文本分类')
win.configure(bg='#a8a8a8')
# win.geometry('1000x650') # 这里的乘号不是 * ，而是小写英文字母 x
screenwidth = win.winfo_screenwidth()  # 屏幕宽度
screenheight = win.winfo_screenheight()  # 屏幕高度
win.geometry('{}x{}+{}+{}'.format(screenwidth, screenheight, 0, 0))  # 大小以及位置


# 整个窗口加一个背景页
wrap_frm = tkinter.Frame(master=win,background='#a8a8a8')
wrap_frm.place(relx=0.1,y=10,relheight=1,relwidth=0.8)

# 左侧板块，输入文本
left_frm = tkinter.Frame(master=wrap_frm,background='#a8a8a8')
left_frm.place(relx=0.05,y=0,height=600,relwidth=0.3)

# 右侧板块，显示文本
right_frm = tkinter.Frame(master=wrap_frm,background='#a8a8a8')
right_frm.place(relx=0.4,y=0,relheight=1,relwidth=0.55)

# tkinter.Label(frm,text="1111").pack(side=tkinter.LEFT)

'''窗口头部
# head_lb = Label(win,text='新闻文本分类',\
#         bg='#d3fbfb',\
#         fg='red',\
#         font=('华文新魏',32),\
#         width=20,\
#         height=1,\
#         relief=FLAT)#(平的)SUNKEN)
# head_lb.place(relx=0,y=0,height=0.4,width=200)
# # head_lb.pack()'''

# 批量输入新闻的区域
batch_input_frm = tkinter.Frame(master=left_frm,background='#BBBBBB',height=300)
batch_input_frm.pack(fill=tkinter.X,ipady=0)

batch_input = tkinter.Label(batch_input_frm,text='1.批量输入',\
        bg='#DCDCDC',\
        fg='red',\
        font=('华文新魏',32),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
batch_input.pack()

file_input_frm = tkinter.Frame(master=batch_input_frm,background='#BBBBBB',height=50)
file_input_frm.pack(fill=tkinter.X)

file_input_entry_var = tkinter.StringVar()
file_input_path = tkinter.Entry(file_input_frm, width=20,textvariable=file_input_entry_var)
file_input_path.place(relx=0.28,y=10,height=30,relwidth=0.7)

filePath = ''
def selectFile():
    global filePath
    filePath = tkinter.filedialog.askopenfilename(title=u'选择文件', filetypes=(("files", "*.csv;*.xlsx"),("All files", "*.*")))
    file_input_entry_var.set(os.path.basename(filePath))
    file_input_path.configure(state="readonly")#tkinter.DISABLED
    # print(os.path.basename(filePath))
    print(filePath, os.path.splitext(filePath))
file_input = tkinter.Button(file_input_frm, text='选择文件', font=('华文新魏',15), command = selectFile)
file_input.place(x=0,y=10,height=30,relwidth=0.25)

# 预测进度显示
def progress_bar_run(str):
    global progress_bar
    progress_bar.configure(text=str)

# 预测进度显示
progress_bar_frm = tkinter.Frame(master=batch_input_frm,background='#BBBBBB',height=40)
progress_bar_frm.pack(fill=tkinter.X)
progress_bar = tkinter.Label(progress_bar_frm,\
        text='预测进度',\
        bg='#FFFFFF',\
        fg='blue',\
        font=('华文新魏',20),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
progress_bar.place(x=0,y=0,height=40,relwidth=1)

# 保存新闻及预测结果
complete_predict_line = list()
# 当前训练的条数
current_num = 0
# 总的新闻文章数量
news_total_num = 0
# 批量提交进行预测
def batch_commit_run():
    global forecastdataLen
    forecastdataLen = connectMySQL.forecastdataLen()
    global news_total_num, current_num
    current_num = 0
    news_total_num = 0
    print('批量预测')

    # 完整预测条目：[编号, channel, title, content]
    global complete_predict_line

    # 读取文件 filePath
    file_data = xlrd.open_workbook(filePath)
    sheet_names = file_data.sheet_names()
    # sheet = file_data.sheet_by_name("类别")
    sheet = file_data.sheet_by_index(0)

    # 对列名单独处理
    columns_name = sheet.row_values(0)
    columns_name.append('预测时长')
    complete_predict_line.append(columns_name)
    # 记录预测条数
    # num_i = 0
    news_total_num = sheet.nrows - 1
    # 忽略第一行列名
    for i in range(1, sheet.nrows):
        # print(i)
        current_num = i
        line = sheet.row_values(i)
        text = ""
        # 拼接标题和内容后放入模型预测
        for item in line[2:]:
            text += item
            # print(text)

        #记录预测时间
        start_time = time.clock()
        # 预测
        predict_label = prediction_fun(text)
        end_time = time.clock()
        cost_time = end_time - start_time
        # print("预测第{}条".format(current_num),cost_time)
        # print("<" + str(line[0]) + ">" + line[2])

        # 存放完整条目
        complete_predict_line.append([int(line[0]), predict_label, line[2], line[3], "{:.4f}秒".format(cost_time)])

        # 写入数据库
        sql = "insert into forecastdata values ({},\"{}\",\"{}\",\"{}\",{})".format(current_num+forecastdataLen, escape_string(predict_label), escape_string(line[2]), escape_string(line[3]), cost_time)
        # sql = "insert into forecastdata values (\"{}\",\"{}\",\"{}\",{})".format(escape_string(predict_label), escape_string(line[2]), escape_string(line[3]), cost_time)
        connectMySQL.insertData(sql)

        # cost_time = "{:.4f}秒".format(cost_time)  # end_time - start_time)
        # 进度条效果
        percentage = current_num * 100 / news_total_num
        text='预测进度：{:.2f}%'.format(percentage)
        progress_bar_run(text)  # 创建一个线程实例
        progress_bar.update()
        global current_page_num
        if current_num==1:
            news_all_update(1)
            current_page_num = 1
        # progress_bar_run(text)

batch_commit = tkinter.Button(master=batch_input_frm, text='开始预测',font=('华文新魏',15), command=batch_commit_run)
batch_commit.pack(fill=tkinter.X,pady=5 )

def file_save_run():
    global forecastdataLen
    global filePath
    if filePath != '':
        # 保存结果
        external_name = os.path.splitext(filePath)[-1]
        file_save_path = tkinter.filedialog.asksaveasfilename(title=u'保存文件', filetypes=(("files", external_name),("All files", "*.*"))) + external_name
        myWorkbook = openpyxl.Workbook()
        mySheet = myWorkbook.active
        mySheet.title = "预测结果"
        columns_name = (('编号', 'channelName', 'title', 'content', '预测时长(秒)'),)

        sql = " select * from forecastdata where id > {}".format(forecastdataLen)
        itemData = connectMySQL.selectAllData(sql)
        itemData = columns_name + itemData
        # print(itemData)

        for k in range(len(itemData)):
            for j in range(len(itemData[k])):
                mySheet.cell(k + 1, j + 1, value=str(itemData[k][j]))
        myWorkbook.save(file_save_path)
        print('保存文件')

    # for k in range(len(complete_predict_line)):
    #     for j in range(len(complete_predict_line[k])):
    #         mySheet.cell(k + 1, j + 1, value=str(complete_predict_line[k][j]))
    # myWorkbook.save(file_save_path)
    # # return item_list
    # print('保存文件')

file_save = tkinter.Button(master=batch_input_frm, text='保存结果',font=('华文新魏',15), command=file_save_run)
file_save.pack(fill=tkinter.X)

tkinter.Frame(master=left_frm,background='#BBBBBB',height=20).pack(fill=tkinter.X)

single_input_frm = tkinter.Frame(master=left_frm,background='#BBBBBB',height=200)
single_input_frm.pack(fill=tkinter.X)

single_input = tkinter.Label(single_input_frm,text='2.单例输入',\
        bg='#DCDCDC',\
        fg='red',\
        font=('华文新魏',32),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
single_input.pack()

title_input_frm = tkinter.Frame(master=single_input_frm,background='#BBBBBB',height=40)
title_input_frm.pack(fill=tkinter.X)

title_input = tkinter.Label(title_input_frm,text='新闻标题:',\
        bg='#BBBBBB',\
        # fg='red',\
        font=('华文新魏',15),\
        width=60,\
        height=1,\
        relief=tkinter.FLAT)
title_input.place(x=0,y=5,height=30,relwidth=0.25)

entry_var = tkinter.StringVar()
title_input_entry = tkinter.Entry(title_input_frm, width=20,textvariable=entry_var)
title_input_entry.place(relx=0.28,y=5,height=30,relwidth=0.7)
# entry_var.set("我是一个Entry")
# print(title_input_entry.get())

content_input_frm = tkinter.Frame(master=single_input_frm,background='#BBBBBB',height=40)
content_input_frm.pack(fill=tkinter.X)
content_input = tkinter.Label(content_input_frm,text='新闻内容:',\
        bg='#BBBBBB',\
        # fg='red',\
        font=('华文新魏',15),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
content_input.place(x=0,y=5,height=35,relwidth=0.25)

content_entry_var = tkinter.StringVar()
content_input_entry = tkinter.Entry(content_input_frm, width=20,textvariable=content_entry_var)
content_input_entry.place(relx=0.28,y=5,height=30,relwidth=0.7)
# content_entry_var.set("我是一个内容Entry")
# print(content_input_entry.get())

def single_commit_run():
    global  news_total_num, current_num, current_page_num, forecastdataLen, filePath
    filePath=''
    forecastdataLen = connectMySQL.forecastdataLen()
    current_num = 1
    news_total_num = 1
    text = title_input_entry.get() + content_input_entry.get()
    # 记录预测时间
    start_time = time.clock()
    # 预测
    predict_label = prediction_fun(text)
    end_time = time.clock()
    cost_time = "{:.4f}秒".format(end_time - start_time)
    class_name.configure(text='预测分类：{}'.format(predict_label))
    class_time.configure(text='预测时长：{}'.format(cost_time))
    news_title.configure(text=title_input_entry.get())
    news_content.delete(1.0, tkinter.END)
    news_content.insert(tkinter.END, content_input_entry.get())
    current_page.configure(text='当前第{}页'.format(1))
    current_page_num = 1
    right_frm.update()

    # 写入数据库
    sql = "insert into forecastdata values ({},\"{}\",\"{}\",\"{}\",{})".format(1+forecastdataLen, escape_string(predict_label), escape_string(title_input_entry.get()), escape_string(content_input_entry.get()), end_time - start_time)
    # sql = "insert into forecastdata values (\"{}\",\"{}\",\"{}\",{})".format(escape_string(predict_label), escape_string(line[2]), escape_string(line[3]), cost_time)
    connectMySQL.insertData(sql)

    print('单例预测')

single_commit = tkinter.Button(master=single_input_frm, text='开始预测',font=('华文新魏',15), command=single_commit_run)
single_commit.pack(fill=tkinter.X)


def news_all_update(page_num):
    global forecastdataLen
    # 查数据库
    sql = " select * from forecastdata where id = {} ".format(page_num+forecastdataLen)
    itemData = connectMySQL.selectOneData(sql)
    # print(itemData)

    # 更新页面
    class_name.configure(text='预测分类：{}'.format(itemData[1]))
    class_time.configure(text='预测时长：{}秒'.format(itemData[4]))
    news_title.configure(text=itemData[2])
    news_content.delete(1.0, tkinter.END)
    news_content.insert(tkinter.END, itemData[3])
    current_page.configure(text='当前第{}页,共{}页'.format(page_num, news_total_num))
    right_frm.update()

    # 更新页面
    # class_name.configure(text='预测分类：{}'.format(complete_predict_line[page_num][1]))
    # class_time.configure(text='预测时长：{}'.format(complete_predict_line[page_num][4]))
    # news_title.configure(text=complete_predict_line[page_num][2])
    # news_content.delete(1.0, tkinter.END)
    # news_content.insert(tkinter.END, complete_predict_line[page_num][3])
    # current_page.configure(text='当前第{}页,共{}页'.format(page_num, news_total_num))
    # right_frm.update()

#新闻文本显示区域
predict_frm = tkinter.Frame(master=right_frm,height=35)
predict_frm.pack(fill=tkinter.X,padx=5)
class_name = tkinter.Label(predict_frm,text='预测分类：名称',\
        # bg='#A9A9A9',\
        # fg='red',\
        font=('华文新魏',13),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
class_name.place(relx=0,y=2.5,height=30,width=140)
# class_name.configure(text = '11')
class_time = tkinter.Label(predict_frm,text='预测时长：时长',\
        # bg='#A9A9A9',\
        # fg='red',\
        font=('华文新魏',13),\
        width=200,\
        height=1,\
        relief=tkinter.FLAT)
class_time.place(x=150,y=2.5,height=30,width=200)

news_frm = tkinter.Frame(master=right_frm,background='#FFFFFF',height=500)
news_frm.pack(fill=tkinter.X, padx=5)

# print(news_frm.width)Message
news_title = tkinter.Message(news_frm,text='新闻标题',\
        bg='#FFFFFF',\
        # fg='red',\
        font=('华文新魏',18),\
        width=screenwidth*0.5*0.8,\
        # height=1,\
        relief=tkinter.FLAT)
news_title.pack(fill=tkinter.X, pady=1)


tkinter.Frame(master=news_frm,background='#CCCCCC',height=5).place(height=400) #pack(fill=tkinter.X)

news_content=tkinter.Text(news_frm,font=('华文新魏',14),autoseparators=False,height=27)
news_content.pack(fill=tkinter.X)#place(relx=0.05,y=43,height=457,relwidth=0.9)
news_content.delete(1.0, tkinter.END)
news_content.insert(tkinter.END,'   新闻内容')

page_frm = tkinter.Frame(master=right_frm,height=150)
page_frm.pack(fill=tkinter.X,padx=5)

def front_page_run():
    global current_page_num
    if current_page_num == 1:
        pass
    elif current_page_num == 0:
        pass
    else:
        current_page_num -= 1
        news_all_update(current_page_num)
        print('前一页')

front_page = tkinter.Button(master=page_frm, font=('华文新魏',15),text='前一页', command=front_page_run)
front_page.pack(side=tkinter.LEFT)

def current_page_run():
    print('当前页')

# tkinter.Button(master=page_frm, text='当前页', command=current_page_run)
current_page_num = 0
current_page = tkinter.Label(page_frm,text='当前第0页',\
        # bg='#FFB666',\
        fg='blue',\
        font=('华文新魏',15),\
        width=20,\
        height=1,\
        relief=tkinter.FLAT)
current_page.pack(side=tkinter.LEFT,expand=tkinter.YES,fill=tkinter.Y)

def behind_page_run():
    global current_page_num
    global current_num
    global news_total_num
    # print(current_page_num,current_num,news_total_num)
    if current_page_num == news_total_num:
        pass
    elif current_page_num == current_num:
        pass
    else:
        current_page_num += 1
        news_all_update(current_page_num)
        print('后一页')

behind_page = tkinter.Button(master=page_frm,text='后一页',font=('华文新魏',15), command=behind_page_run)
behind_page.pack(side=tkinter.LEFT)

win.mainloop()
